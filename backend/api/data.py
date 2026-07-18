"""
数据浏览 API
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import os
from database import get_db
from models import ScrapedItem, Task

router = APIRouter()


@router.get("/tasks/{task_id}/items", summary="获取任务采集结果")
async def get_task_items(
    task_id: int,
    platform: Optional[str] = None,
    item_type: Optional[str] = None,
    sort_by: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
):
    query = db.query(ScrapedItem).filter(ScrapedItem.task_id == task_id)
    if platform:
        query = query.filter(ScrapedItem.platform == platform)
    if item_type:
        query = query.filter(ScrapedItem.item_type == item_type)

    if sort_by == "price_asc":
        query = query.order_by(ScrapedItem.price.asc())
    elif sort_by == "price_desc":
        query = query.order_by(ScrapedItem.price.desc())
    elif sort_by == "likes":
        query = query.order_by(ScrapedItem.likes.desc())
    elif sort_by == "sales":
        query = query.order_by(ScrapedItem.sales.desc())
    else:
        query = query.order_by(ScrapedItem.id.desc())

    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [_item_to_dict(item) for item in items],
    }


@router.get("/tasks/{task_id}/stats", summary="获取任务统计概览")
async def get_task_stats(task_id: int, db: Session = Depends(get_db)):
    items = db.query(ScrapedItem).filter(ScrapedItem.task_id == task_id).all()
    platforms = {}
    total_price = 0
    price_count = 0
    for item in items:
        p = item.platform
        if p not in platforms:
            platforms[p] = {"count": 0, "platform": p}
        platforms[p]["count"] += 1
        if item.price and item.price > 0:
            total_price += item.price
            price_count += 1
    return {
        "total_items": len(items),
        "by_platform": list(platforms.values()),
        "avg_price": round(total_price / price_count, 2) if price_count > 0 else 0,
    }


@router.get("/items/{item_id}", summary="获取单条数据详情")
async def get_item_detail(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ScrapedItem).filter(ScrapedItem.id == item_id).first()
    if not item:
        raise HTTPException(404, "数据不存在")
    return _item_to_dict(item)


@router.post("/tasks/{task_id}/export", summary="导出为 Excel")
async def export_excel(task_id: int, db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(404, "任务不存在")

    items = db.query(ScrapedItem).filter(ScrapedItem.task_id == task_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{task.keyword}"

    if task.task_type == "product_group":
        headers = ["序号", "平台", "标题", "价格(¥)", "原价", "销量", "店铺", "评分", "主图URL", "链接"]
        widths = [6, 10, 45, 10, 10, 12, 18, 8, 50, 50]
        ws.append(headers)
        for idx, item in enumerate(items, 1):
            ws.append([
                idx, item.platform, item.title or "", item.price,
                item.original_price, item.sales, item.shop_name or "",
                item.rating, item.main_image or "", item.source_url or "",
            ])
    else:
        headers = ["序号", "平台", "标题", "文案", "点赞", "评论", "收藏", "作者", "发布时间", "标签", "链接"]
        widths = [6, 10, 45, 50, 10, 8, 8, 15, 12, 25, 50]
        ws.append(headers)
        for idx, item in enumerate(items, 1):
            ws.append([
                idx, item.platform, item.title or "", item.content_text or "",
                item.likes, item.comments_count, item.favorites,
                item.author_name or "", item.publish_time or "",
                ", ".join(item.tags) if item.tags else "",
                item.source_url or "",
            ])

    # 列宽
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 冻结首行
    ws.freeze_panes = "A2"

    export_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
    os.makedirs(export_dir, exist_ok=True)
    filename = f"task_{task_id}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return {"filename": filename, "download_url": f"/api/data/download/{filename}"}


@router.get("/download/{filename}", summary="下载 Excel 文件")
async def download_file(filename: str):
    export_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
    filepath = os.path.join(export_dir, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "文件不存在")
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


def _item_to_dict(item: ScrapedItem) -> dict:
    return {
        "id": item.id, "task_id": item.task_id,
        "platform": item.platform, "item_type": item.item_type,
        "title": item.title, "price": item.price,
        "original_price": item.original_price, "sales": item.sales,
        "main_image": item.main_image, "detail_images": item.detail_images,
        "shop_name": item.shop_name, "specs": item.specs, "rating": item.rating,
        "content_text": item.content_text, "cover_image": item.cover_image,
        "likes": item.likes, "comments_count": item.comments_count,
        "favorites": item.favorites, "publish_time": item.publish_time,
        "author_name": item.author_name, "tags": item.tags,
        "comments": item.comments, "source_url": item.source_url,
        "created_at": str(item.created_at) if item.created_at else None,
    }
